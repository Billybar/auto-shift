import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from ortools.sat.python import cp_model


def solve_shift_scheduling():
    # -------------------------------------------------------------------------
    # 1. Data Definition
    # -------------------------------------------------------------------------
    # ADDED: 'max_mornings' and 'max_evenings' to each employee
    employees = [
        {'name': 'Barak', 'target_shifts': 6, 'max_shifts': 6, 'max_nights': 0, 'max_mornings': 6, 'max_evenings': 0,
         'history_streak': 0},
        {'name': 'Billy', 'target_shifts': 5, 'max_shifts': 5, 'max_nights': 2, 'max_mornings': 3, 'max_evenings': 3,
         'history_streak': 6},
        {'name': 'Asaf', 'target_shifts': 5, 'max_shifts': 6, 'max_nights': 2, 'max_mornings': 0, 'max_evenings': 3,
         'history_streak': 2},
        {'name': 'Gadi', 'target_shifts': 5, 'max_shifts': 6, 'max_nights': 2, 'max_mornings': 3, 'max_evenings': 4,
         'history_streak': 0},
        {'name': 'Saar', 'target_shifts': 5, 'max_shifts': 6, 'max_nights': 2, 'max_mornings': 3, 'max_evenings': 5,
         'history_streak': 0},
        {'name': 'Ira', 'target_shifts': 5, 'max_shifts': 6, 'max_nights': 1, 'max_mornings': 5, 'max_evenings': 5,
         'history_streak': 0},
        {'name': 'Shon', 'target_shifts': 3, 'max_shifts': 4, 'max_nights': 2, 'max_mornings': 2, 'max_evenings': 5,
         'history_streak': 0},
        {'name': 'Gilad', 'target_shifts': 3, 'max_shifts': 4, 'max_nights': 2, 'max_mornings': 2, 'max_evenings': 3,
         'history_streak': 0},
        {'name': 'Dolev', 'target_shifts': 3, 'max_shifts': 4, 'max_nights': 3, 'max_mornings': 0, 'max_evenings': 5,
         'history_streak': 0},
        {'name': 'Michael', 'target_shifts': 3, 'max_shifts': 4, 'max_nights': 2, 'max_mornings': 4, 'max_evenings': 5,
         'history_streak': 0},
    ]

    # Color palette
    employee_colors = [
        'FF9999', '99FF99', '9999FF', 'FFFF99', 'FFCC99',
        'FF99FF', '99FFFF', 'CCCCCC', 'D9D9D9', 'E6B8B7'
    ]

    num_employees = len(employees)
    num_days = 7
    num_shifts = 3  # 0=Morning, 1=Noon, 2=Night
    shifts_per_day_demand = 2

    # Hard Constraints: Unavailability Requests
    unavailable_requests = [
        # ID 0
        (0, 0, 2), (0, 1, 1), (0, 1, 2), (0, 5, 1), (0, 5, 2),
        # ID 1
        (1, 0, 0), (1, 1, 2), (1, 2, 0), (1, 2, 2), (1, 3, 0), (1, 3, 2),
        (1, 4, 1), (1, 4, 2), (1, 5, 1), (1, 5, 2), (1, 6, 0), (1, 6, 1), (1, 6, 2),
        # ID 2
        (2, 4, 1), (2, 5, 1), (2, 5, 2), (2, 6, 0), (2, 6, 1),
        # ID 3
        (3, 1, 0), (3, 1, 1), (3, 2, 1), (3, 3, 2),
        (3, 4, 0), (3, 4, 1), (3, 4, 2), (3, 5, 0),
        # ID 4
        (4, 6, 1),
        # ID 5
        (5, 0, 1),
        # ID 6
        (6, 0, 0), (6, 1, 0), (6, 2, 0), (6, 2, 2), (6, 3, 0), (6, 3, 2),
        (6, 4, 0), (6, 4, 2), (6, 6, 1), (6, 6, 2),
        # ID 7
        (7, 0, 0), (7, 0, 1), (7, 0, 2), (7, 1, 0), (7, 1, 1), (7, 1, 2),
        (7, 2, 2), (7, 3, 1), (7, 3, 2), (7, 4, 1), (7, 4, 2), (7, 5, 0),
        (7, 6, 1), (7, 6, 2),
        # ID 8
        (8, 0, 0), (8, 4, 2), (8, 5, 0), (8, 5, 1), (8, 5, 2), (8, 6, 0),
        # ID 9
        (9, 0, 0), (9, 1, 0), (9, 1, 1), (9, 1, 2), (9, 2, 0), (9, 2, 1),
        (9, 3, 0), (9, 4, 0), (9, 4, 1), (9, 4, 2), (9, 5, 0),
    ]

    model = cp_model.CpModel()

    # -------------------------------------------------------------------------
    # 2. Variables
    # -------------------------------------------------------------------------
    shift_vars = {}
    for e in range(num_employees):
        for d in range(num_days):
            for s in range(num_shifts):
                shift_vars[(e, d, s)] = model.NewBoolVar(f'shift_{e}_{d}_{s}')

    # -------------------------------------------------------------------------
    # 3. Hard Constraints
    # -------------------------------------------------------------------------

    # A. Exact number of workers per shift
    for d in range(num_days):
        for s in range(num_shifts):
            model.Add(sum(shift_vars[(e, d, s)] for e in range(num_employees)) == shifts_per_day_demand)

    # B. Prevent back-to-back shifts
    for e in range(num_employees):
        for total_s in range(num_days * num_shifts - 1):
            day = total_s // num_shifts
            shift = total_s % num_shifts
            next_total_s = total_s + 1
            next_day = next_total_s // num_shifts
            next_shift = next_total_s % num_shifts
            model.Add(shift_vars[(e, day, shift)] + shift_vars[(e, next_day, next_shift)] <= 1)

    # C. Unavailability
    for req in unavailable_requests:
        e, d, s = req
        model.Add(shift_vars[(e, d, s)] == 0)

    # D. Prevent working 7 days in a row
    for e in range(num_employees):
        work_days_vars = []
        for d in range(num_days):
            is_working_day = model.NewBoolVar(f'working_day_{e}_{d}')
            model.Add(sum(shift_vars[(e, d, s)] for s in range(num_shifts)) > 0).OnlyEnforceIf(is_working_day)
            model.Add(sum(shift_vars[(e, d, s)] for s in range(num_shifts)) == 0).OnlyEnforceIf(is_working_day.Not())
            work_days_vars.append(is_working_day)

        streak = employees[e]['history_streak']
        if streak > 0:
            limit = 7 - streak
            if limit <= num_days and limit > 0:
                model.Add(sum(work_days_vars[0:limit]) < limit)
        if streak == 0:
            model.Add(sum(work_days_vars) < 7)

    # -------------------------------------------------------------------------
    # 4. Soft Constraints (Optimization)
    # -------------------------------------------------------------------------
    WEIGHT_NIGHTS = 5
    WEIGHT_MORNINGS = 4  # NEW
    WEIGHT_EVENINGS = 4  # NEW
    WEIGHT_REST_GAP = 2
    WEIGHT_TARGET_SHIFTS = 4
    WEIGHT_CONSECUTIVE_NIGHTS = 20  # NEW - High penalty for 3 nights in a row

    objective_terms = []

    for e in range(num_employees):
        emp_shifts = []
        morning_shifts = []
        evening_shifts = []
        night_shifts = []

        for d in range(num_days):
            # Collect shift variables by type
            morning_shifts.append(shift_vars[(e, d, 0)])
            evening_shifts.append(shift_vars[(e, d, 1)])
            night_shifts.append(shift_vars[(e, d, 2)])

            for s in range(num_shifts):
                emp_shifts.append(shift_vars[(e, d, s)])

        # 1. Limit Max Night Shifts
        excess_nights = model.NewIntVar(0, 7, f'excess_nights_{e}')
        model.Add(sum(night_shifts) <= employees[e]['max_nights'] + excess_nights)
        objective_terms.append(excess_nights * WEIGHT_NIGHTS)

        # 2. Limit Max Morning Shifts (NEW)
        excess_mornings = model.NewIntVar(0, 7, f'excess_mornings_{e}')
        model.Add(sum(morning_shifts) <= employees[e]['max_mornings'] + excess_mornings)
        objective_terms.append(excess_mornings * WEIGHT_MORNINGS)

        # 3. Limit Max Evening Shifts (NEW)
        excess_evenings = model.NewIntVar(0, 7, f'excess_evenings_{e}')
        model.Add(sum(evening_shifts) <= employees[e]['max_evenings'] + excess_evenings)
        objective_terms.append(excess_evenings * WEIGHT_EVENINGS)

        # 4. Avoid 3 Consecutive Nights (NEW)
        # Checking windows of 3 days: [d, d+1, d+2]
        for d in range(num_days - 2):
            is_three_nights = model.NewBoolVar(f'3nights_{e}_{d}')
            # Logic: IF night[d] AND night[d+1] AND night[d+2] THEN is_three_nights = True
            model.AddBoolAnd([
                shift_vars[(e, d, 2)],
                shift_vars[(e, d + 1, 2)],
                shift_vars[(e, d + 2, 2)]
            ]).OnlyEnforceIf(is_three_nights)

            # Helper to force it to 0 if not 3 nights (for strict correctness, though optimization handles it)
            model.AddBoolOr([
                shift_vars[(e, d, 2)].Not(),
                shift_vars[(e, d + 1, 2)].Not(),
                shift_vars[(e, d + 2, 2)].Not()
            ]).OnlyEnforceIf(is_three_nights.Not())

            objective_terms.append(is_three_nights * WEIGHT_CONSECUTIVE_NIGHTS)

        # 5. Avoid Short Rest Gaps (8-8)
        for total_s in range(num_days * num_shifts - 2):
            day = total_s // num_shifts
            shift = total_s % num_shifts
            target_total_s = total_s + 2
            t_day = target_total_s // num_shifts
            t_shift = target_total_s % num_shifts

            both_working = model.NewBoolVar(f'bad_gap_{e}_{total_s}')
            model.AddBoolAnd([shift_vars[(e, day, shift)], shift_vars[(e, t_day, t_shift)]]).OnlyEnforceIf(both_working)
            model.AddBoolOr([shift_vars[(e, day, shift)].Not(), shift_vars[(e, t_day, t_shift)].Not()]).OnlyEnforceIf(
                both_working.Not())
            objective_terms.append(both_working * WEIGHT_REST_GAP)

        # 6. Target Shifts
        total_worked = sum(emp_shifts)
        delta = model.NewIntVar(0, 21, f'delta_target_{e}')
        model.Add(total_worked - employees[e]['target_shifts'] <= delta)
        model.Add(employees[e]['target_shifts'] - total_worked <= delta)
        objective_terms.append(delta * WEIGHT_TARGET_SHIFTS)

        # Max shifts limit
        model.Add(total_worked <= employees[e]['max_shifts'])

    # -------------------------------------------------------------------------
    # 5. Solve and Export
    # -------------------------------------------------------------------------
    model.Minimize(sum(objective_terms))
    solver = cp_model.CpSolver()
    status = solver.Solve(model)

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        print(f"Solution Found! Total Penalty Cost: {solver.ObjectiveValue()}")
        create_excel_schedule(solver, shift_vars, employees, num_days, num_shifts, shifts_per_day_demand,
                              employee_colors)
    else:
        print("No feasible solution found with the current constraints.")


def create_excel_schedule(solver, shift_vars, employees, num_days, num_shifts, shifts_per_day_demand, colors):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.sheet_view.rightToLeft = True

    days_names = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]
    shifts_names = ["בוקר", "צהריים", "לילה"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Header Row
    ws.cell(row=1, column=1).value = "משמרת"
    for i, day in enumerate(days_names):
        cell = ws.cell(row=1, column=i + 2)
        cell.value = day
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    current_row = 2
    for s_idx, s_name in enumerate(shifts_names):
        for slot in range(shifts_per_day_demand):
            shift_cell = ws.cell(row=current_row, column=1)
            shift_cell.value = f"{s_name} ({slot + 1})"
            shift_cell.font = Font(bold=True)
            shift_cell.alignment = center_align
            shift_cell.border = thin_border

            for d in range(num_days):
                assigned_workers = []
                for e_idx, emp in enumerate(employees):
                    if solver.Value(shift_vars[(e_idx, d, s_idx)]):
                        assigned_workers.append(e_idx)

                cell = ws.cell(row=current_row, column=d + 2)
                cell.border = thin_border
                cell.alignment = center_align

                if len(assigned_workers) > slot:
                    worker_idx = assigned_workers[slot]
                    worker_name = employees[worker_idx]['name']
                    worker_color = colors[worker_idx]

                    cell.value = worker_name
                    cell.fill = PatternFill(start_color=worker_color, end_color=worker_color, fill_type="solid")
            current_row += 1
        current_row += 1

    # Summary Table
    summary_row = current_row + 2
    ws.cell(row=summary_row, column=1).value = "סיכום עובדים"
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=14)

    headers = ["שם", "סהכ משמרות", "לילות", "בקרים", "ערבים"]
    for col_idx, h in enumerate(headers):
        ws.cell(row=summary_row + 1, column=col_idx + 1).value = h
        ws.cell(row=summary_row + 1, column=col_idx + 1).font = Font(bold=True)

    for i, emp in enumerate(employees):
        r = summary_row + 2 + i
        name_cell = ws.cell(row=r, column=1)
        name_cell.value = emp['name']
        name_cell.fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")

        total = 0
        nights = 0
        mornings = 0
        evenings = 0

        for d in range(num_days):
            if solver.Value(shift_vars[(i, d, 0)]): mornings += 1
            if solver.Value(shift_vars[(i, d, 1)]): evenings += 1
            if solver.Value(shift_vars[(i, d, 2)]): nights += 1

        total = mornings + evenings + nights

        ws.cell(row=r, column=2).value = total
        ws.cell(row=r, column=3).value = nights
        ws.cell(row=r, column=4).value = mornings
        ws.cell(row=r, column=5).value = evenings

    wb.save("shift_schedule_colored.xlsx")
    print("Excel file created successfully.")


if __name__ == "__main__":
    solve_shift_scheduling()