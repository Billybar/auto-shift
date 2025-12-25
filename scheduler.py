import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from ortools.sat.python import cp_model

# ==========================================
#         System Configuration
# ==========================================

# Enable image parsing?
# True = Try to read constraints from the image file
# False = Ignore image, use only the manual list below
ENABLE_IMAGE_PARSING = False

# Image filename (if enabled)
IMAGE_FILENAME = "images/image3.png"


# ==========================================

def solve_shift_scheduling():
    # -------------------------------------------------------------------------
    # 1. Data Definition
    # -------------------------------------------------------------------------
    employees = [
        {'name': 'Ira', 'target_shifts': 5, 'max_shifts': 6,
         'max_nights': 1, 'min_nights': 1,
         'max_mornings': 3, 'min_mornings': 2,
         'max_evenings': 3, 'min_evenings': 0,
         'history_streak': 2},

        {'name': 'Asaf', 'target_shifts': 5, 'max_shifts': 6,
         'max_nights': 2, 'min_nights': 2,
         'max_mornings': 0, 'min_mornings': 0,
         'max_evenings': 3, 'min_evenings': 3,
         'history_streak': 1},

        {'name': 'Barak', 'target_shifts': 5, 'max_shifts': 6,
         'max_nights': 0, 'min_nights': 0,
         'max_mornings': 6, 'min_mornings': 6,
         'max_evenings': 0, 'min_evenings': 0,
         'history_streak': 4},

        {'name': 'Gilad', 'target_shifts': 3, 'max_shifts': 4,
         'max_nights': 1, 'min_nights': 1,
         'max_mornings': 3, 'min_mornings': 0,
         'max_evenings': 4, 'min_evenings': 0,
         'history_streak': 0},

        {'name': 'Gadi', 'target_shifts': 5, 'max_shifts': 6,
         'max_nights': 2, 'min_nights': 2,
         'max_mornings': 3, 'min_mornings': 0,
         'max_evenings': 3, 'min_evenings': 5,
         'history_streak': 3},

        {'name': 'Dolev', 'target_shifts': 4, 'max_shifts': 4,
         'max_nights': 3, 'min_nights': 2,
         'max_mornings': 2, 'min_mornings': 0,
         'max_evenings': 4, 'min_evenings': 0,
         'history_streak': 1},

        {'name': 'Michael', 'target_shifts': 3, 'max_shifts': 4,
         'max_nights': 2, 'min_nights': 0,
         'max_mornings': 2, 'min_mornings': 0,
         'max_evenings': 2, 'min_evenings': 0,
         'history_streak': 0},

        {'name': 'Saar', 'target_shifts': 4, 'max_shifts': 4,
         'max_nights': 2, 'min_nights': 1,
         'max_mornings': 3, 'min_mornings': 0,
         'max_evenings': 5, 'min_evenings': 0,
         'history_streak': 5},

        {'name': 'Billy', 'target_shifts': 5, 'max_shifts': 5,
         'max_nights': 2, 'min_nights': 1,
         'max_mornings': 3, 'min_mornings': 2,
         'max_evenings': 3, 'min_evenings': 3,
         'history_streak': 0},

        {'name': 'Shon', 'target_shifts': 3, 'max_shifts': 4,
         'max_nights': 2, 'min_nights': 0,
         'max_mornings': 2, 'min_mornings': 0,
         'max_evenings': 4, 'min_evenings': 0,
         'history_streak': 0},
    ]

    # Previous Week Context
    # Enter the IDs of employees who worked last Saturday

    worked_last_sat_noon = [0, 4]  # Example: Ira and Gadi
    worked_last_sat_night = [1,5]  # Example: Asaf, Dolev

    employee_colors = [
        'FF9999', '99FF99', '9999FF', 'FFFF99', 'FFCC99',
        'FF99FF', '99FFFF', 'CCCCCC', '87CEFA', 'E6B8B7'
    ]

    num_employees = len(employees)
    num_days = 7
    num_shifts = 3
    shifts_per_day_demand = 2

    # --------------------------------------------------------
    # Manual Constraints List
    # --------------------------------------------------------
    # Enter manual constraints here
    # Format: (Employee ID, Day 0-6, Shift 0-2)
    manual_requests = [
        # --- ID 0: Ira ---
        (0, 3, 0), (0, 3, 1), (0, 3, 2),  # Wednesday: Full Day
        (0, 4, 0), (0, 4, 1), (0, 4, 2),  # Thursday: Full Day

        # --- ID 1: Asaf ---
        (1, 5, 1), (1, 5, 2),  # Friday: Afternoon, Night
        (1, 6, 0), (1, 6, 1),  # Saturday: Morning, Afternoon

        # --- ID 2: Barak ---
        (2, 5, 0),  # Friday: Morning
        (2, 6, 0),  # Saturday: Morning

        # --- ID 3: Gilad ---
        (3, 0, 0), (3, 0, 1), (3, 0, 2),  # Sunday: Full Day
        (3, 1, 0), (3, 1, 1), (3, 1, 2),  # Monday: Full Day
        (3, 2, 2),  # Tuesday: Night
        (3, 3, 1), (3, 3, 2),  # Wednesday: Afternoon, Night
        (3, 4, 2),  # Thursday: Night
        (3, 6, 1), (3, 6, 2),  # Saturday: Afternoon, Night

        # --- ID 4: Gadi ---
        (4, 0, 0), (4, 0, 1), (4, 0, 2),  # Sunday: Full Day
        (4, 3, 1), (4, 3, 2),  # Wednesday: Afternoon, Night

        # --- ID 5: Dolev ---
        (5, 4, 0), (5, 4, 1), (5, 4, 2),  # Thursday: Full Day
        (5, 5, 0), (5, 5, 1), (5, 5, 2),  # Friday: Full Day

        # --- ID 6: Michael ---
        (6, 0, 0), (6, 0, 1), (6, 0, 2),  # Sunday: Full Day
        (6, 1, 2),  # Monday: Night
        (6, 2, 2),  # Tuesday: Night
        (6, 3, 2),  # Wednesday: Night
        (6, 4, 2),  # Thursday: Night

        # ID 7

        # --- ID 8: Billy ---
        (8, 0, 0), (8, 0, 1),  # Sunday: Morning, Afternoon
        (8, 1, 2),  # Monday: Night
        (8, 2, 0), (8, 2, 2),  # Tuesday: Morning, Night
        (8, 3, 0),  # Wednesday: Morning
        (8, 4, 1), (8, 4, 2),  # Thursday: Afternoon, Night
        (8, 5, 1), (8, 5, 2),  # Friday: Afternoon, Night
        (8, 6, 0), (8, 6, 1), (8, 6, 2),  # Saturday: Full Day

        # ID 9
    ]

    # --------------------------------------------------------
    # Manual Assignments (Force Shift)
    # --------------------------------------------------------
    # These employees MUST be assigned to these specific shifts.
    manual_assignments = [
        # Example: Force Employee 0 (Ira) to work Sunday Morning
        # (0, 0, 0),

        # Example: Force Employee 2 (Barak) to work Tuesday Night
        # (2, 2, 2),
    ]

    # --------------------------------------------------------
    # Image Parsing Logic (Only if enabled)
    # --------------------------------------------------------
    image_constraints = []

    if ENABLE_IMAGE_PARSING:
        print(f"--- Image Mode Enabled: Parsing {IMAGE_FILENAME} ---")
        if os.path.exists(IMAGE_FILENAME):
            try:
                # Import only if needed to avoid crashes if file is missing/broken
                from image_process.cv2_image_parser import ScheduleImageParser

                # Define employee order in image (Top -> Down)
                img_employee_order = [0, 1, 2, 3, 4, 5, 6, 7]

                parser = ScheduleImageParser(IMAGE_FILENAME)
                # Call the new parse_tables function
                image_constraints = parser.parse_tables(img_employee_order)
                print(f"V Success: Extracted {len(image_constraints)} constraints from image.")
            except ImportError:
                print("X Error: cv2_image_parser.py is missing or invalid.")
            except AttributeError:
                print("X Error: 'parse_tables' function not found in cv2_image_parser.py. Check file version.")
            except Exception as e:
                print(f"X General Error parsing image: {e}")
        else:
            print(f"X Error: File {IMAGE_FILENAME} not found.")
    else:
        print("--- Image Mode Disabled: Using manual list only ---")

    # Combine Lists
    unavailable_requests = manual_requests + image_constraints

    # Debug Print
    print(f"\nTotal active constraints: {len(unavailable_requests)}")
    days_map = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    shifts_map = ["Morning", "Noon", "Night"]

    if len(unavailable_requests) > 0:
        print("Constraints Details:")
        for req in unavailable_requests:
            e, d, s = req
            if 0 <= e < len(employees):
                print(f"- {employees[e]['name']}: {days_map[d]} {shifts_map[s]}")

    model = cp_model.CpModel()

    # -------------------------------------------------------------------------
    # 2. Variables
    # -------------------------------------------------------------------------
    shift_vars = {}
    for e in range(num_employees):
        for d in range(num_days):
            for s in range(num_shifts):
                shift_vars[(e, d, s)] = model.NewBoolVar(f'shift_{e}_{d}_{s}')

    # -------------------------------------------------------------------------
    # 3. Hard Constraints
    # -------------------------------------------------------------------------

    # A. Exact number of workers per shift
    for d in range(num_days):
        for s in range(num_shifts):
            model.Add(sum(shift_vars[(e, d, s)] for e in range(num_employees)) == shifts_per_day_demand)

    # B. Prevent back-to-back shifts
    for e in range(num_employees):
        for total_s in range(num_days * num_shifts - 1):
            day = total_s // num_shifts
            shift = total_s % num_shifts
            next_total_s = total_s + 1
            next_day = next_total_s // num_shifts
            next_shift = next_total_s % num_shifts
            model.Add(shift_vars[(e, day, shift)] + shift_vars[(e, next_day, next_shift)] <= 1)

    # C. Apply unavailability (Manual + Image)
    for req in unavailable_requests:
        e, d, s = req
        if 0 <= e < num_employees and 0 <= d < num_days and 0 <= s < num_shifts:
            model.Add(shift_vars[(e, d, s)] == 0)

    # D. Apply Previous Week Constraints (Rest Rules)
    # Employees who worked last Saturday Night cannot work Sunday Morning.
    for emp_id in worked_last_sat_night:
        if 0 <= emp_id < num_employees:
            model.Add(shift_vars[(emp_id, 0, 0)] == 0)

    # E. Apply Manual Assignments (Force Specific Shifts)
    for assign in manual_assignments:
        e, d, s = assign
        # Avoid indexError
        if 0 <= e < num_employees and 0 <= d < num_days and 0 <= s < num_shifts:
            # Validate against unavailability constraints
            if (e, d, s) in unavailable_requests:
                emp_name = employees[e]['name']
                raise ValueError(f"CRITICAL ERROR: Conflict detected for {emp_name}! "
                                 f"You are forcing a shift (Day {d}, Shift {s}) but it is marked as UNAVAILABLE.")

            print(f"Forcing assignment: Employee {e} -> Day {d} Shift {s}")
            model.Add(shift_vars[(e, d, s)] == 1)


    # F. Prevent working 7 days in a row
    for e in range(num_employees):
        work_days_vars = []
        for d in range(num_days):
            is_working_day = model.NewBoolVar(f'working_day_{e}_{d}')
            model.Add(sum(shift_vars[(e, d, s)] for s in range(num_shifts)) > 0).OnlyEnforceIf(is_working_day)
            model.Add(sum(shift_vars[(e, d, s)] for s in range(num_shifts)) == 0).OnlyEnforceIf(is_working_day.Not())
            work_days_vars.append(is_working_day)

        streak = employees[e]['history_streak']
        if streak > 0:
            limit = 7 - streak
            if limit <= num_days and limit > 0:
                model.Add(sum(work_days_vars[0:limit]) < limit)
        if streak == 0:
            model.Add(sum(work_days_vars) < 7)

    # -------------------------------------------------------------------------
    # 4. Soft Constraints (Optimization)
    # -------------------------------------------------------------------------
    WEIGHT_TARGET_SHIFTS = 4
    WEIGHT_REST_GAP = 2

    WEIGHT_MAX_NIGHTS = 5
    WEIGHT_MAX_MORNINGS = 4
    WEIGHT_MAX_EVENINGS = 4
    WEIGHT_CONSECUTIVE_NIGHTS = 20

    WEIGHT_MIN_NIGHTS = 5
    WEIGHT_MIN_MORNINGS = 4
    WEIGHT_MIN_EVENINGS = 4

    objective_terms = []

    for e in range(num_employees):
        emp_shifts = []
        morning_shifts = []
        evening_shifts = []
        night_shifts = []

        for d in range(num_days):
            morning_shifts.append(shift_vars[(e, d, 0)])
            evening_shifts.append(shift_vars[(e, d, 1)])
            night_shifts.append(shift_vars[(e, d, 2)])

            for s in range(num_shifts):
                emp_shifts.append(shift_vars[(e, d, s)])

        # Maximum Constraints
        excess_nights = model.NewIntVar(0, 7, f'excess_nights_{e}')
        model.Add(sum(night_shifts) <= employees[e]['max_nights'] + excess_nights)
        objective_terms.append(excess_nights * WEIGHT_MAX_NIGHTS)

        excess_mornings = model.NewIntVar(0, 7, f'excess_mornings_{e}')
        model.Add(sum(morning_shifts) <= employees[e]['max_mornings'] + excess_mornings)
        objective_terms.append(excess_mornings * WEIGHT_MAX_MORNINGS)

        excess_evenings = model.NewIntVar(0, 7, f'excess_evenings_{e}')
        model.Add(sum(evening_shifts) <= employees[e]['max_evenings'] + excess_evenings)
        objective_terms.append(excess_evenings * WEIGHT_MAX_EVENINGS)

        # Minimum Constraints
        shortage_nights = model.NewIntVar(0, 7, f'shortage_nights_{e}')
        model.Add(sum(night_shifts) + shortage_nights >= employees[e]['min_nights'])
        objective_terms.append(shortage_nights * WEIGHT_MIN_NIGHTS)

        shortage_mornings = model.NewIntVar(0, 7, f'shortage_mornings_{e}')
        model.Add(sum(morning_shifts) + shortage_mornings >= employees[e]['min_mornings'])
        objective_terms.append(shortage_mornings * WEIGHT_MIN_MORNINGS)

        shortage_evenings = model.NewIntVar(0, 7, f'shortage_evenings_{e}')
        model.Add(sum(evening_shifts) + shortage_evenings >= employees[e]['min_evenings'])
        objective_terms.append(shortage_evenings * WEIGHT_MIN_EVENINGS)

        # Logic and Rest
        # Avoid 3 consecutive nights
        for d in range(num_days - 2):
            is_three_nights = model.NewBoolVar(f'3nights_{e}_{d}')
            model.AddBoolAnd(
                [shift_vars[(e, d, 2)], shift_vars[(e, d + 1, 2)], shift_vars[(e, d + 2, 2)]]).OnlyEnforceIf(
                is_three_nights)
            model.AddBoolOr([shift_vars[(e, d, 2)].Not(), shift_vars[(e, d + 1, 2)].Not(),
                             shift_vars[(e, d + 2, 2)].Not()]).OnlyEnforceIf(is_three_nights.Not())
            objective_terms.append(is_three_nights * WEIGHT_CONSECUTIVE_NIGHTS)

        # Avoid short rest gap
        for total_s in range(num_days * num_shifts - 2):
            day = total_s // num_shifts
            shift = total_s % num_shifts
            t_total_s = total_s + 2
            t_day = t_total_s // num_shifts
            t_shift = t_total_s % num_shifts
            both_working = model.NewBoolVar(f'bad_gap_{e}_{total_s}')
            model.AddBoolAnd([shift_vars[(e, day, shift)], shift_vars[(e, t_day, t_shift)]]).OnlyEnforceIf(both_working)
            model.AddBoolOr([shift_vars[(e, day, shift)].Not(), shift_vars[(e, t_day, t_shift)].Not()]).OnlyEnforceIf(
                both_working.Not())
            objective_terms.append(both_working * WEIGHT_REST_GAP)

        # include Previous Week Gap (saturday noon and night)
        if e in worked_last_sat_noon:
            objective_terms.append(shift_vars[(e, 0, 0)] * WEIGHT_REST_GAP)

        # Optimization: If worked Last Sat Night, avoid Sun Noon.
        if e in worked_last_sat_night:
            objective_terms.append(shift_vars[(e, 0, 1)] * WEIGHT_REST_GAP)


        # Target Shifts
        total_worked = sum(emp_shifts)
        delta = model.NewIntVar(0, 21, f'delta_target_{e}')
        model.Add(total_worked - employees[e]['target_shifts'] <= delta)
        model.Add(employees[e]['target_shifts'] - total_worked <= delta)
        objective_terms.append(delta * WEIGHT_TARGET_SHIFTS)

        # Hard limit max shifts
        model.Add(total_worked <= employees[e]['max_shifts'])

    # -------------------------------------------------------------------------
    # 5. Solve and Export
    # -------------------------------------------------------------------------
    model.Minimize(sum(objective_terms))
    solver = cp_model.CpSolver()
    status = solver.Solve(model)

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        print(f"\n✅ Solution Found! Cost (Penalty): {solver.ObjectiveValue()}")
        create_excel_schedule(solver, shift_vars, employees, num_days, num_shifts, shifts_per_day_demand,
                              employee_colors)
    else:
        print("\n❌ No feasible solution found. Try relaxing constraints.")


def create_excel_schedule(solver, shift_vars, employees, num_days, num_shifts, shifts_per_day_demand, colors):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.sheet_view.rightToLeft = True

    days_names = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]
    shifts_names = ["בוקר", "צהריים", "לילה"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Header Row
    ws.cell(row=1, column=1).value = "משמרת"
    for i, day in enumerate(days_names):
        cell = ws.cell(row=1, column=i + 2)
        cell.value = day
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    current_row = 2
    for s_idx, s_name in enumerate(shifts_names):
        for slot in range(shifts_per_day_demand):
            shift_cell = ws.cell(row=current_row, column=1)
            shift_cell.value = f"{s_name} ({slot + 1})"
            shift_cell.font = Font(bold=True)
            shift_cell.alignment = center_align
            shift_cell.border = thin_border

            for d in range(num_days):
                assigned_workers = []
                for e_idx, emp in enumerate(employees):
                    if solver.Value(shift_vars[(e_idx, d, s_idx)]):
                        assigned_workers.append(e_idx)

                cell = ws.cell(row=current_row, column=d + 2)
                cell.border = thin_border
                cell.alignment = center_align

                if len(assigned_workers) > slot:
                    worker_idx = assigned_workers[slot]
                    worker_name = employees[worker_idx]['name']
                    worker_color = colors[worker_idx]

                    cell.value = worker_name
                    cell.fill = PatternFill(start_color=worker_color, end_color=worker_color, fill_type="solid")
            current_row += 1
        current_row += 1

    # Summary Table
    summary_row = current_row + 2
    ws.cell(row=summary_row, column=1).value = "סיכום עובדים"
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=14)

    headers = ["שם", "סהכ משמרות", "לילות", "בקרים", "ערבים"]
    for col_idx, h in enumerate(headers):
        ws.cell(row=summary_row + 1, column=col_idx + 1).value = h
        ws.cell(row=summary_row + 1, column=col_idx + 1).font = Font(bold=True)

    for i, emp in enumerate(employees):
        r = summary_row + 2 + i
        name_cell = ws.cell(row=r, column=1)
        name_cell.value = emp['name']
        name_cell.fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")

        total = 0
        nights = 0
        mornings = 0
        evenings = 0

        for d in range(num_days):
            if solver.Value(shift_vars[(i, d, 0)]): mornings += 1
            if solver.Value(shift_vars[(i, d, 1)]): evenings += 1
            if solver.Value(shift_vars[(i, d, 2)]): nights += 1

        total = mornings + evenings + nights

        ws.cell(row=r, column=2).value = total
        ws.cell(row=r, column=3).value = nights
        ws.cell(row=r, column=4).value = mornings
        ws.cell(row=r, column=5).value = evenings

    wb.save("shift_schedule_colored.xlsx")
    print("Excel file created successfully: shift_schedule_colored.xlsx")


if __name__ == "__main__":
    solve_shift_scheduling()