# excel_writer.py
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def create_excel_schedule(solver, shift_vars, employees, num_days, num_shifts, shifts_per_day_demand, colors):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.sheet_view.rightToLeft = True

    days_names = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]
    shifts_names = ["בוקר", "צהריים", "לילה"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Header Row
    ws.cell(row=1, column=1).value = "משמרת"
    for i, day in enumerate(days_names):
        cell = ws.cell(row=1, column=i + 2)
        cell.value = day
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    current_row = 2
    for s_idx, s_name in enumerate(shifts_names):
        for slot in range(shifts_per_day_demand):
            shift_cell = ws.cell(row=current_row, column=1)
            shift_cell.value = f"{s_name} ({slot + 1})"
            shift_cell.font = Font(bold=True)
            shift_cell.alignment = center_align
            shift_cell.border = thin_border

            for d in range(num_days):
                assigned_workers = []
                for e_idx, emp in enumerate(employees):
                    if solver.Value(shift_vars[(e_idx, d, s_idx)]):
                        assigned_workers.append(e_idx)

                cell = ws.cell(row=current_row, column=d + 2)
                cell.border = thin_border
                cell.alignment = center_align

                if len(assigned_workers) > slot:
                    worker_idx = assigned_workers[slot]
                    worker_name = employees[worker_idx]['name']
                    worker_color = colors[worker_idx]

                    cell.value = worker_name
                    cell.fill = PatternFill(start_color=worker_color, end_color=worker_color, fill_type="solid")
            current_row += 1
        current_row += 1

    # Summary Table
    summary_row = current_row + 2
    ws.cell(row=summary_row, column=1).value = "סיכום עובדים"
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=14)

    headers = ["שם", "סהכ משמרות", "לילות", "בקרים", "ערבים"]
    for col_idx, h in enumerate(headers):
        ws.cell(row=summary_row + 1, column=col_idx + 1).value = h
        ws.cell(row=summary_row + 1, column=col_idx + 1).font = Font(bold=True)

    for i, emp in enumerate(employees):
        r = summary_row + 2 + i
        name_cell = ws.cell(row=r, column=1)
        name_cell.value = emp['name']
        name_cell.fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")

        total = 0
        nights = 0
        mornings = 0
        evenings = 0

        for d in range(num_days):
            if solver.Value(shift_vars[(i, d, 0)]): mornings += 1
            if solver.Value(shift_vars[(i, d, 1)]): evenings += 1
            if solver.Value(shift_vars[(i, d, 2)]): nights += 1

        total = mornings + evenings + nights

        ws.cell(row=r, column=2).value = total
        ws.cell(row=r, column=3).value = nights
        ws.cell(row=r, column=4).value = mornings
        ws.cell(row=r, column=5).value = evenings

    wb.save("shift_schedule_output/shift_schedule_colored.xlsx")
    print("Excel file created successfully: shift_schedule_colored.xlsx")